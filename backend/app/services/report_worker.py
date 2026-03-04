"""
BeeYield Report Generation Worker
===================================
Generates high-fidelity PDF and Excel reports using real data.
Uses ReportLab for PDF layout and openpyxl for Excel.
"""
import io
import os
import uuid
import math
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Any

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, cm
from reportlab.lib.colors import HexColor, white, black
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, PageBreak, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.db.supabase_db import db_select, db_insert, db_update


# ============ COLOR PALETTE ============
BRAND_GREEN = HexColor("#1B9157")
BRAND_DARK = HexColor("#0F172A")
BRAND_YELLOW = HexColor("#F4D03F")
BRAND_LIGHT = HexColor("#F8FAFC")
BRAND_GRAY = HexColor("#64748B")


# ============ DATA AGGREGATOR ============
class DataAggregator:
    """Queries multiple tables and computes KPIs for reports."""

    def __init__(self, user_id: str, token: Optional[str] = None):
        self.user_id = user_id
        self.token = token

    async def get_apiaries(self, place_id: Optional[str] = None) -> List[Dict]:
        filters = {"user_id": self.user_id}
        if place_id:
            filters["id"] = place_id
        return await db_select("apiaries", filters=filters, token=self.token)

    async def get_hives(self, apiary_id: Optional[str] = None) -> List[Dict]:
        filters = {"user_id": self.user_id}
        if apiary_id:
            filters["apiary_id"] = apiary_id
        return await db_select("hives", filters=filters, token=self.token)

    async def get_harvests(self, scope_days: int = 365, apiary_id: Optional[str] = None) -> List[Dict]:
        since = (datetime.now() - timedelta(days=scope_days)).isoformat()
        filters = {"user_id": self.user_id}
        if apiary_id:
            filters["apiary_id"] = apiary_id
        harvests = await db_select("harvests", filters=filters, order_by="harvest_date", ascending=False, token=self.token)
        return [h for h in harvests if h.get("harvest_date", "") >= since[:10]]

    async def get_billing(self, scope_days: int = 365) -> List[Dict]:
        since = (datetime.now() - timedelta(days=scope_days)).isoformat()
        filters = {"user_id": self.user_id}
        ledger = await db_select("billing_ledger", filters=filters, order_by="date", ascending=False, token=self.token)
        return [b for b in ledger if b.get("date", "") >= since]

    async def get_inspections(self, scope_days: int = 90) -> List[Dict]:
        since = (datetime.now() - timedelta(days=scope_days)).isoformat()
        filters = {"user_id": self.user_id}
        inspections = await db_select("inspections", filters=filters, order_by="inspection_date", ascending=False, token=self.token)
        return [i for i in inspections if i.get("inspection_date", "") >= since[:10]]

    async def compute_kpis(self, scope_days: int = 365) -> Dict[str, Any]:
        apiaries = await self.get_apiaries()
        hives = await self.get_hives()
        harvests = await self.get_harvests(scope_days)
        billing = await self.get_billing(scope_days)

        total_harvest_kg = sum(float(h.get("quantity_kg", 0)) for h in harvests)
        active_hives = [h for h in hives if h.get("status", "").upper() in ("ACTIVE", "HEALTHY", "GOOD")]
        avg_yield = total_harvest_kg / len(active_hives) if active_hives else 0

        income = sum(float(b.get("amount", 0)) for b in billing if b.get("transaction_type") == "income")
        expenses = sum(float(b.get("amount", 0)) for b in billing if b.get("transaction_type") == "expense")

        return {
            "total_apiaries": len(apiaries),
            "total_hives": len(hives),
            "active_hives": len(active_hives),
            "total_harvest_kg": round(total_harvest_kg, 2),
            "avg_yield_per_hive_kg": round(avg_yield, 2),
            "harvest_count": len(harvests),
            "total_income": round(income, 2),
            "total_expenses": round(expenses, 2),
            "net_revenue": round(income - expenses, 2),
            "scope_days": scope_days,
            "generated_at": datetime.now().isoformat()
        }


# ============ PDF GENERATOR ============
class PDFReportGenerator:
    """Generates branded PDF reports using ReportLab."""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    def _setup_styles(self):
        self.styles.add(ParagraphStyle(
            'BrandTitle', parent=self.styles['Title'],
            fontName='Helvetica-Bold', fontSize=28, textColor=BRAND_DARK,
            spaceAfter=6 * mm
        ))
        self.styles.add(ParagraphStyle(
            'BrandSubtitle', parent=self.styles['Normal'],
            fontName='Helvetica', fontSize=11, textColor=BRAND_GRAY,
            spaceAfter=12 * mm
        ))
        self.styles.add(ParagraphStyle(
            'SectionHeader', parent=self.styles['Heading2'],
            fontName='Helvetica-Bold', fontSize=16, textColor=BRAND_GREEN,
            spaceBefore=8 * mm, spaceAfter=4 * mm
        ))
        self.styles.add(ParagraphStyle(
            'KPIValue', fontName='Helvetica-Bold', fontSize=22,
            textColor=BRAND_DARK, alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            'KPILabel', fontName='Helvetica', fontSize=9,
            textColor=BRAND_GRAY, alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            'BodyText', parent=self.styles['Normal'],
            fontName='Helvetica', fontSize=10, textColor=BRAND_DARK,
            leading=14
        ))

    def generate(self, kpis: Dict, apiaries: List, hives: List,
                 harvests: List, billing: List, inspections: List,
                 sections: List[str], user_name: str = "Beekeeper") -> bytes:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=20 * mm, rightMargin=20 * mm,
            topMargin=25 * mm, bottomMargin=20 * mm
        )

        story = []

        # === COVER ===
        story.append(Spacer(1, 30 * mm))
        story.append(Paragraph("BEEYIELD", self.styles['BrandTitle']))
        story.append(Paragraph(
            f"Season Summary Report • {datetime.now().strftime('%B %Y')}",
            self.styles['BrandSubtitle']
        ))
        story.append(HRFlowable(
            width="100%", thickness=2, color=BRAND_GREEN,
            spaceAfter=8 * mm, spaceBefore=4 * mm
        ))
        story.append(Paragraph(
            f"Prepared for: <b>{user_name}</b><br/>"
            f"Period: Last {kpis['scope_days']} days<br/>"
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            self.styles['BodyText']
        ))
        story.append(Spacer(1, 15 * mm))

        # === KPI DASHBOARD ===
        story.append(Paragraph("Key Performance Indicators", self.styles['SectionHeader']))
        kpi_data = [
            [self._kpi_cell("Total Harvest", f"{kpis['total_harvest_kg']} kg"),
             self._kpi_cell("Avg Yield/Hive", f"{kpis['avg_yield_per_hive_kg']} kg"),
             self._kpi_cell("Active Hives", str(kpis['active_hives']))],
            [self._kpi_cell("Total Revenue", f"KES {kpis['total_income']:,.0f}"),
             self._kpi_cell("Expenses", f"KES {kpis['total_expenses']:,.0f}"),
             self._kpi_cell("Net Profit", f"KES {kpis['net_revenue']:,.0f}")]
        ]
        kpi_table = Table(kpi_data, colWidths=[56 * mm, 56 * mm, 56 * mm])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), BRAND_LIGHT),
            ('BOX', (0, 0), (-1, -1), 0.5, BRAND_GRAY),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, HexColor("#E2E8F0")),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 10 * mm))

        # === APIARIES ===
        if 'apiaries' in sections and apiaries:
            story.append(Paragraph("Apiaries Overview", self.styles['SectionHeader']))
            apiary_header = ['Name', 'Location', 'Hives', 'Status', 'Forage']
            apiary_rows = [apiary_header]
            for a in apiaries[:20]:
                apiary_rows.append([
                    a.get("name", "—"),
                    a.get("location_name", a.get("region", "—")),
                    str(a.get("hive_count", 0)),
                    a.get("status", "active").capitalize(),
                    a.get("primary_forage", a.get("forage_type", "—"))
                ])
            t = Table(apiary_rows, colWidths=[40*mm, 35*mm, 20*mm, 25*mm, 35*mm])
            t.setStyle(self._table_style(len(apiary_rows)))
            story.append(t)
            story.append(Spacer(1, 8 * mm))

        # === HIVES ===
        if 'hives' in sections and hives:
            story.append(Paragraph("Hive Inventory", self.styles['SectionHeader']))
            hive_header = ['Code', 'Type', 'Status', 'Health']
            hive_rows = [hive_header]
            for h in hives[:30]:
                hive_rows.append([
                    h.get("hive_code", "—"),
                    h.get("hive_type", "—"),
                    h.get("status", "—"),
                    h.get("health_status", "—")
                ])
            t = Table(hive_rows, colWidths=[40*mm, 35*mm, 35*mm, 40*mm])
            t.setStyle(self._table_style(len(hive_rows)))
            story.append(t)
            story.append(Spacer(1, 8 * mm))

        # === HARVESTS ===
        if 'harvests' in sections and harvests:
            story.append(Paragraph("Harvest Log", self.styles['SectionHeader']))
            harvest_header = ['Date', 'Qty (kg)', 'Type', 'Batch', 'Verified']
            harvest_rows = [harvest_header]
            for h in harvests[:30]:
                harvest_rows.append([
                    str(h.get("harvest_date", "—"))[:10],
                    str(h.get("quantity_kg", 0)),
                    h.get("honey_type", "—"),
                    h.get("batch_code", "—"),
                    "✓" if h.get("is_verified") else "—"
                ])
            t = Table(harvest_rows, colWidths=[30*mm, 25*mm, 30*mm, 40*mm, 20*mm])
            t.setStyle(self._table_style(len(harvest_rows)))
            story.append(t)
            story.append(Spacer(1, 8 * mm))

        # === INSPECTIONS ===
        if 'inspections' in sections and inspections:
            story.append(Paragraph("Recent Inspections", self.styles['SectionHeader']))
            insp_header = ['Date', 'Hive', 'Result', 'Notes']
            insp_rows = [insp_header]
            for i in inspections[:15]:
                insp_rows.append([
                    str(i.get("inspection_date", "—"))[:10],
                    i.get("hive_id", "—")[:8] + "...",
                    i.get("overall_health", i.get("result", "—")),
                    (i.get("notes", "") or "")[:40]
                ])
            t = Table(insp_rows, colWidths=[28*mm, 28*mm, 30*mm, 60*mm])
            t.setStyle(self._table_style(len(insp_rows)))
            story.append(t)
            story.append(Spacer(1, 8 * mm))

        # === FOOTER ===
        story.append(Spacer(1, 15 * mm))
        story.append(HRFlowable(width="100%", thickness=1, color=BRAND_GRAY))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(
            f"This report was automatically generated by BeeYield on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}. "
            "Data is sourced from your live BeeYield dashboard. For questions, contact support@beeyield.com.",
            ParagraphStyle('Footer', fontName='Helvetica', fontSize=8, textColor=BRAND_GRAY)
        ))

        doc.build(story)
        return buffer.getvalue()

    def _kpi_cell(self, label: str, value: str) -> Paragraph:
        return Paragraph(
            f'<font size="18"><b>{value}</b></font><br/>'
            f'<font size="8" color="#64748B">{label}</font>',
            ParagraphStyle('kpi', alignment=TA_CENTER, leading=22, fontName='Helvetica-Bold')
        )

    def _table_style(self, num_rows: int) -> TableStyle:
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), BRAND_GREEN),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, BRAND_LIGHT]),
            ('BOX', (0, 0), (-1, -1), 0.5, BRAND_GRAY),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, HexColor("#E2E8F0")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ])


# ============ EXCEL GENERATOR ============
class ExcelReportGenerator:
    """Generates branded Excel reports using openpyxl."""

    HEADER_FILL = PatternFill(start_color="1B9157", end_color="1B9157", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    DATA_FONT = Font(name="Calibri", size=10)
    KPI_FONT = Font(name="Calibri", bold=True, size=14)
    BORDER = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )

    def generate(self, kpis: Dict, apiaries: List, hives: List,
                 harvests: List, billing: List, inspections: List,
                 sections: List[str], user_name: str = "Beekeeper") -> bytes:
        wb = openpyxl.Workbook()

        # Summary Sheet
        ws = wb.active
        ws.title = "Summary"
        ws.append(["BeeYield Report"])
        ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="1B9157")
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([f"User: {user_name}"])
        ws.append([f"Period: Last {kpis['scope_days']} days"])
        ws.append([])

        kpi_items = [
            ("Total Harvest (kg)", kpis['total_harvest_kg']),
            ("Avg Yield/Hive (kg)", kpis['avg_yield_per_hive_kg']),
            ("Active Hives", kpis['active_hives']),
            ("Total Hives", kpis['total_hives']),
            ("Total Apiaries", kpis['total_apiaries']),
            ("Total Income (KES)", kpis['total_income']),
            ("Total Expenses (KES)", kpis['total_expenses']),
            ("Net Revenue (KES)", kpis['net_revenue']),
        ]
        for label, val in kpi_items:
            ws.append([label, val])
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18

        # Apiaries Sheet
        if 'apiaries' in sections and apiaries:
            ws2 = wb.create_sheet("Apiaries")
            headers = ['Name', 'Location', 'Hives', 'Status', 'Forage', 'Established']
            self._write_header(ws2, headers)
            for a in apiaries:
                ws2.append([
                    a.get("name", ""), a.get("location_name", ""),
                    a.get("hive_count", 0), a.get("status", ""),
                    a.get("primary_forage", ""), str(a.get("established_date", ""))[:10]
                ])
            self._auto_width(ws2)

        # Hives Sheet
        if 'hives' in sections and hives:
            ws3 = wb.create_sheet("Hives")
            headers = ['Code', 'Type', 'Status', 'Health']
            self._write_header(ws3, headers)
            for h in hives:
                ws3.append([
                    h.get("hive_code", ""), h.get("hive_type", ""),
                    h.get("status", ""), h.get("health_status", "")
                ])
            self._auto_width(ws3)

        # Harvests Sheet
        if 'harvests' in sections and harvests:
            ws4 = wb.create_sheet("Harvests")
            headers = ['Date', 'Quantity (kg)', 'Type', 'Batch Code', 'Moisture %', 'Color', 'Verified']
            self._write_header(ws4, headers)
            for h in harvests:
                ws4.append([
                    str(h.get("harvest_date", ""))[:10],
                    float(h.get("quantity_kg", 0)),
                    h.get("honey_type", ""),
                    h.get("batch_code", ""),
                    float(h.get("moisture_content_percent", 0)),
                    h.get("color_grade", ""),
                    "Yes" if h.get("is_verified") else "No"
                ])
            self._auto_width(ws4)

        # Financial Sheet
        if billing:
            ws5 = wb.create_sheet("Financial")
            headers = ['Date', 'Type', 'Module', 'Description', 'Amount (KES)', 'eTIMS Status']
            self._write_header(ws5, headers)
            for b in billing:
                ws5.append([
                    str(b.get("date", ""))[:10],
                    b.get("transaction_type", ""),
                    b.get("module_type", ""),
                    b.get("description", ""),
                    float(b.get("amount", 0)),
                    b.get("etims_status", "")
                ])
            self._auto_width(ws5)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _write_header(self, ws, headers: List[str]):
        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(self, ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


# ============ ORCHESTRATOR ============
async def generate_report_async(
    user_id: str,
    report_type: str,
    parameters: Dict,
    file_format: str = "PDF",
    token: Optional[str] = None
) -> Dict[str, Any]:
    """
    Main orchestrator: aggregates data, generates file, uploads to Supabase Storage,
    and records the report in the generated_reports table.
    """
    job_id = str(uuid.uuid4())

    # Record pending job
    await db_insert("generated_reports", {
        "id": job_id,
        "user_id": user_id,
        "report_type": report_type,
        "file_format": file_format,
        "parameters": parameters,
        "status": "processing",
        "file_url": None
    }, token=token)

    try:
        scope_days = int(parameters.get("scope_days", 365))
        sections = parameters.get("sections", ["apiaries", "hives", "harvests", "overview"])
        place_id = parameters.get("place_id")
        user_name = parameters.get("user_name", "Beekeeper")

        # Aggregate data
        agg = DataAggregator(user_id, token)
        kpis = await agg.compute_kpis(scope_days)
        apiaries = await agg.get_apiaries(place_id)
        hives = await agg.get_hives()
        harvests = await agg.get_harvests(scope_days)
        billing = await agg.get_billing(scope_days)
        inspections = await agg.get_inspections(scope_days)

        # Generate file
        if file_format.upper() == "XLSX":
            gen = ExcelReportGenerator()
            file_bytes = gen.generate(kpis, apiaries, hives, harvests, billing, inspections, sections, user_name)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"
        else:
            gen = PDFReportGenerator()
            file_bytes = gen.generate(kpis, apiaries, hives, harvests, billing, inspections, sections, user_name)
            content_type = "application/pdf"
            ext = "pdf"

        filename = f"beeyield_report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"

        # Upload to Supabase Storage
        file_url = await _upload_to_storage(file_bytes, filename, content_type, token)

        # Update record
        await db_update(
            "generated_reports",
            {"status": "completed", "file_url": file_url, "file_name": filename},
            {"id": job_id},
            token=token
        )

        return {"job_id": job_id, "status": "completed", "file_url": file_url, "file_name": filename}

    except Exception as e:
        print(f"[REPORT ERROR] {str(e)}")
        await db_update(
            "generated_reports",
            {"status": "failed", "file_url": None},
            {"id": job_id},
            token=token
        )
        return {"job_id": job_id, "status": "failed", "error": str(e)}


async def _upload_to_storage(file_bytes: bytes, filename: str, content_type: str, token: Optional[str] = None) -> str:
    """Upload report file to Supabase Storage and return a public/signed URL."""
    import httpx
    from app.core.config import settings

    bucket = "reports"
    path = f"{filename}"
    url = f"{settings.SUPABASE_URL}/storage/v1/object/{bucket}/{path}"

    headers = {
        "apikey": settings.SUPABASE_KEY,
        "Authorization": f"Bearer {settings.SUPABASE_SERVICE_ROLE_KEY or settings.SUPABASE_KEY}",
        "Content-Type": content_type,
        "x-upsert": "true"
    }

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            resp = await client.post(url, content=file_bytes, headers=headers)
            if resp.status_code in (200, 201):
                public_url = f"{settings.SUPABASE_URL}/storage/v1/object/public/{bucket}/{path}"
                return public_url
            else:
                print(f"[STORAGE] Upload failed {resp.status_code}: {resp.text}")
                # Fallback: return a data URI or placeholder
                return f"/api/v1/reports/download/{filename}"
    except Exception as e:
        print(f"[STORAGE] Upload error: {e}")
        return f"/api/v1/reports/download/{filename}"
